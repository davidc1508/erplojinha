from __future__ import annotations

import json
import re
import unicodedata
import uuid
from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PLANILHAS_DIR = ROOT / "Planilhas"
SEED_FILE = ROOT / "src" / "Lojinha.Infrastructure" / "Migrations" / "CatalogSeedSql.g.cs"
IGNORE_WORKBOOKS: set[str] = set()
CATEGORIES_PATTERN = re.compile(r'private const string CATEGORIES_JSON = """(.*?)""";', re.S)
PRODUCTS_PATTERN = re.compile(r'private const string PRODUCTS_JSON = """(.*?)""";', re.S)
PRINTERS_PATTERN = re.compile(r'private const string PRINTERS_JSON = """(.*?)""";', re.S)
FILAMENTS_PATTERN = re.compile(r'private const string FILAMENTS_JSON = """(.*?)""";', re.S)
MARKETPLACES_PATTERN = re.compile(r'private const string MARKETPLACES_JSON = """(.*?)""";', re.S)
SUPPLIES_PATTERN = re.compile(r'private const string SUPPLIES_JSON = """(.*?)""";', re.S)
SEED_TIMESTAMP = "2026-04-19T00:00:00Z"
DEPRECIATION_DIVISOR = Decimal("2058.333333")
STOCK_QUANTITY = Decimal("3000")
MINIMUM_STOCK = Decimal("500")
CATEGORY_COLORS = ("#f5b2c5", "#f8d8c7", "#98d9d0", "#f8e58c", "#b8e296", "#d5b3ff")


@dataclass(frozen=True)
class WorkbookContext:
    category_name: str
    wholesale_markup: Decimal
    retail_markup: Decimal
    labor_hours: Decimal
    labor_cost_per_hour: Decimal
    additional_costs: Decimal
    printers: dict[str, dict[str, Decimal | str]]
    filaments: dict[str, dict[str, Decimal | str]]
    marketplaces: dict[str, dict[str, Decimal | str]]


def to_decimal(value: object | None) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


def round_decimal(value: Decimal, digits: str = "0.01") -> Decimal:
    return value.quantize(Decimal(digits), rounding=ROUND_HALF_UP)


def parse_time_to_hours(value: object | None) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if hasattr(value, "total_seconds"):
        return Decimal(str(value.total_seconds())) / Decimal("3600")
    if hasattr(value, "hour") and hasattr(value, "minute") and hasattr(value, "second"):
        total_seconds = value.hour * 3600 + value.minute * 60 + value.second
        return Decimal(total_seconds) / Decimal("3600")
    text = str(value)
    if ":" in text:
        hours, minutes, seconds = [int(part) for part in text.split(":")]
        total_seconds = hours * 3600 + minutes * 60 + seconds
        return Decimal(total_seconds) / Decimal("3600")
    return to_decimal(value)


def get_wear_level(usage_level: object | None) -> Decimal:
    normalized = str(usage_level or "").strip().lower()
    return {
        "basico": Decimal("0.10"),
        "medio": Decimal("0.20"),
        "profissional": Decimal("0.30"),
    }.get(normalized, Decimal("0.45"))


def deterministic_uuid(kind: str, name: str) -> str:
    return str(uuid.uuid5(uuid.NAMESPACE_DNS, f"lojinha-{kind}-{name.strip().lower()}"))


def slugify(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^A-Za-z0-9]+", "-", normalized).strip("-")
    return slug.upper() or "ITEM"


def build_product_sku(category_name: str, product_name: str) -> str:
    return f"{slugify(category_name)}-{slugify(product_name)}"


def normalize_product_name(value: object) -> str:
    return str(value).strip()


def get_items_per_plate(sheet, row: int) -> int:
    pieces = int(to_decimal(sheet[f"C{row}"].value or 1))
    return max(1, pieces)


def build_category(category_name: str, position: int) -> dict:
    return {
        "id": deterministic_uuid("category", category_name),
        "name": category_name,
        "description": f"Categoria extraida de {category_name}",
        "colorHex": CATEGORY_COLORS[position % len(CATEGORY_COLORS)],
        "createdAtUtc": SEED_TIMESTAMP,
        "updatedAtUtc": SEED_TIMESTAMP,
    }


def build_product_seed(category_name: str, product_name: str) -> dict:
    identity = f"{category_name}-{product_name}"
    return {
        "id": deterministic_uuid("product", identity),
        "recipeId": deterministic_uuid("product-recipe", identity),
        "recipeItemId": deterministic_uuid("product-recipe-item", identity),
        "name": normalize_product_name(product_name),
        "sku": build_product_sku(category_name, product_name),
        "description": f"Produto extraido da planilha {category_name}",
        "categoryName": category_name,
        "currentStock": 0.0,
        "minimumStock": 2.0,
        "createdAtUtc": SEED_TIMESTAMP,
        "updatedAtUtc": SEED_TIMESTAMP,
    }


def load_seed_json(text: str, pattern: re.Pattern[str]) -> list[dict]:
    match = pattern.search(text)
    if not match:
        raise RuntimeError(f"Nao foi possivel localizar o bloco {pattern.pattern}")
    return json.loads(match.group(1))


def build_workbook_context(path: Path) -> WorkbookContext:
    workbook = load_workbook(path, data_only=True)
    list_sheet = workbook["Listas"]
    product_sheet = workbook["Produto 1"]

    printers: dict[str, dict[str, Decimal | str]] = {}
    for row in range(6, list_sheet.max_row + 1):
        name = list_sheet[f"AK{row}"].value
        if not name:
            continue
        printers[str(name)] = {
            "brand": str(list_sheet[f"AL{row}"].value or ""),
            "return_months": to_decimal(list_sheet[f"AM{row}"].value),
            "machine_cost": to_decimal(list_sheet[f"AN{row}"].value),
            "work_hours_per_day": to_decimal(list_sheet[f"AO{row}"].value),
            "working_days_per_month": to_decimal(list_sheet[f"AP{row}"].value),
            "power_kw": to_decimal(list_sheet[f"AQ{row}"].value),
            "usage_level": str(list_sheet[f"AR{row}"].value or ""),
            "failure_rate": to_decimal(list_sheet[f"AS{row}"].value),
        }

    filaments: dict[str, dict[str, Decimal | str]] = {}
    for row in range(6, list_sheet.max_row + 1):
        name = list_sheet[f"AU{row}"].value
        if not name:
            continue
        filaments[str(name)] = {
            "brand": str(list_sheet[f"AV{row}"].value or ""),
            "description": str(list_sheet[f"AW{row}"].value or ""),
            "spool_weight_kg": to_decimal(list_sheet[f"AX{row}"].value),
            "spool_length_meters": to_decimal(list_sheet[f"AY{row}"].value),
            "cost_brl": to_decimal(list_sheet[f"AZ{row}"].value),
        }

    marketplaces: dict[str, dict[str, Decimal | str]] = {}
    for row in range(6, list_sheet.max_row + 1):
        name = list_sheet[f"BB{row}"].value
        if not name:
            continue
        marketplaces[str(name)] = {
            "fixed_fee": to_decimal(list_sheet[f"BC{row}"].value),
            "percentage_fee": to_decimal(list_sheet[f"BD{row}"].value),
        }

    additional_costs = sum((to_decimal(product_sheet[f"C{row}"].value) for row in range(26, 35)), start=Decimal("0"))

    return WorkbookContext(
        category_name=path.stem,
        wholesale_markup=to_decimal(product_sheet["C37"].value),
        retail_markup=to_decimal(product_sheet["C38"].value),
        labor_hours=to_decimal(product_sheet["C22"].value),
        labor_cost_per_hour=to_decimal(product_sheet["C23"].value),
        additional_costs=additional_costs,
        printers=printers,
        filaments=filaments,
        marketplaces=marketplaces,
    )


def recalculate_product(product: dict, row: int, sheet, context: WorkbookContext) -> dict:
    product_name = normalize_product_name(sheet[f"B{row}"].value)
    printer_name = str(sheet[f"E{row}"].value or "")
    filament_name = str(sheet[f"F{row}"].value or "")
    printer = context.printers[printer_name]
    filament = context.filaments[filament_name]

    weight = to_decimal(sheet[f"I{row}"].value)
    finishing_input = to_decimal(sheet[f"K{row}"].value)
    finishing_percentage = finishing_input / Decimal("100") if finishing_input > 1 else finishing_input
    print_hours = parse_time_to_hours(sheet[f"H{row}"].value)
    tariff_per_kwh = to_decimal(sheet[f"J{row}"].value)

    material_cost = Decimal("0")
    if filament["spool_weight_kg"] and weight:
        material_cost = (filament["cost_brl"] / (filament["spool_weight_kg"] * Decimal("1000"))) * weight

    energy_cost = print_hours * printer["power_kw"] * tariff_per_kwh
    annual_hours = printer["work_hours_per_day"] * printer["working_days_per_month"] * Decimal("12")
    maintenance_cost = Decimal("0")
    if annual_hours > 0 and print_hours > 0:
        maintenance_cost = ((printer["machine_cost"] * get_wear_level(printer["usage_level"])) / annual_hours) * print_hours

    failure_cost = material_cost * printer["failure_rate"]
    finishing_cost = material_cost * finishing_percentage
    return_investment_cost = Decimal("0")
    if printer["return_months"] > 0 and printer["work_hours_per_day"] > 0 and printer["working_days_per_month"] > 0:
        return_investment_cost = (
            printer["machine_cost"]
            / (printer["return_months"] * printer["work_hours_per_day"] * printer["working_days_per_month"])
        ) * print_hours
    depreciation_cost = Decimal("0")
    if printer["return_months"] > 0:
        usage_estimated_annual = printer["return_months"] * printer["work_hours_per_day"] * printer["working_days_per_month"]
        depreciation_cost = usage_estimated_annual / DEPRECIATION_DIVISOR

    labor_cost = context.labor_hours * context.labor_cost_per_hour
    items_per_plate = Decimal(get_items_per_plate(sheet, row))
    composition_cost = (
        material_cost
        + maintenance_cost
        + failure_cost
        + finishing_cost
        + return_investment_cost
        + depreciation_cost
        + labor_cost
        + context.additional_costs
    )
    total_cost = composition_cost + energy_cost
    unit_composition_cost = composition_cost / items_per_plate
    unit_total_cost = total_cost / items_per_plate
    suggested_price = round_decimal(unit_total_cost * context.retail_markup)

    updated = dict(product)
    updated["name"] = product_name
    updated["costPrice"] = float(round_decimal(unit_total_cost))
    updated["salePrice"] = float(suggested_price)
    updated["suggestedPrice"] = float(suggested_price)
    updated["desiredMarkup"] = float(context.retail_markup)
    updated["itemsPerPlate"] = int(items_per_plate)
    updated["estimatedPrintTimeMinutes"] = float(round_decimal(print_hours * Decimal("60")))
    updated["estimatedWeightGrams"] = float(round_decimal(weight))
    updated["lengthMetersUsed"] = float(round_decimal(to_decimal(sheet[f"G{row}"].value)))
    updated["tariffPerKwh"] = float(round_decimal(tariff_per_kwh))
    updated["finishingPercentage"] = float(round_decimal(finishing_percentage))
    updated["printerName"] = printer_name
    updated["filamentName"] = filament_name
    updated["marketplaceName"] = "NENHUM"
    updated["compositionCost"] = float(round_decimal(unit_composition_cost))
    updated["laborHours"] = float(round_decimal(context.labor_hours))
    updated["laborCostPerHour"] = float(round_decimal(context.labor_cost_per_hour))
    updated["additionalCosts"] = float(round_decimal(context.additional_costs))
    updated["wholesaleMarkup"] = float(round_decimal(context.wholesale_markup))
    updated["retailMarkup"] = float(round_decimal(context.retail_markup))
    updated["resellerMarkup"] = float(round_decimal(context.retail_markup))
    updated["recipeItemQuantity"] = float(round_decimal(weight))
    updated["updatedAtUtc"] = SEED_TIMESTAMP
    return updated


def replace_json_constant(text: str, constant_name: str, payload: list[dict]) -> str:
    serialized = json.dumps(payload, ensure_ascii=True, separators=(",", ":"))
    pattern = re.compile(rf'(private const string {constant_name} = """)(.*?)(""";)', re.S)
    if not pattern.search(text):
        raise RuntimeError(f"Constante {constant_name} nao encontrada")
    return pattern.sub(lambda match: f"{match.group(1)}{serialized}{match.group(3)}", text, count=1)


def main() -> None:
    text = SEED_FILE.read_text(encoding="utf-8")
    categories = load_seed_json(text, CATEGORIES_PATTERN)
    products = load_seed_json(text, PRODUCTS_PATTERN)
    printers = load_seed_json(text, PRINTERS_PATTERN)
    filaments = load_seed_json(text, FILAMENTS_PATTERN)
    marketplaces = load_seed_json(text, MARKETPLACES_PATTERN)
    supplies = load_seed_json(text, SUPPLIES_PATTERN)

    category_index = {item["name"]: item for item in categories}
    product_index = {(item["categoryName"], normalize_product_name(item["name"])): position for position, item in enumerate(products)}
    printer_index = {item["name"]: item for item in printers}
    filament_index = {item["name"]: item for item in filaments}
    marketplace_index = {item["name"]: item for item in marketplaces}
    supply_index = {item["name"]: item for item in supplies}

    updated_products = 0
    added_products = 0
    normalized_finishing = 0

    for workbook_path in sorted(PLANILHAS_DIR.glob("*.xlsx")):
        if workbook_path.name in IGNORE_WORKBOOKS:
            continue

        context = build_workbook_context(workbook_path)
        if context.category_name not in category_index:
            category_index[context.category_name] = build_category(context.category_name, len(category_index))
        workbook = load_workbook(workbook_path, data_only=True)
        list_sheet = workbook["Listas"]

        for printer_name, printer in context.printers.items():
            if printer_name not in printer_index:
                printer_index[printer_name] = {
                    "id": deterministic_uuid("printer", printer_name),
                    "name": printer_name,
                    "brand": printer["brand"],
                    "returnMonths": float(round_decimal(printer["return_months"])),
                    "machineCost": float(round_decimal(printer["machine_cost"])),
                    "workHoursPerDay": float(round_decimal(printer["work_hours_per_day"])),
                    "workingDaysPerMonth": float(round_decimal(printer["working_days_per_month"])),
                    "powerKw": float(round_decimal(printer["power_kw"])),
                    "usageLevel": printer["usage_level"],
                    "failureRate": float(round_decimal(printer["failure_rate"])),
                    "createdAtUtc": SEED_TIMESTAMP,
                    "updatedAtUtc": SEED_TIMESTAMP,
                }
            else:
                current = printer_index[printer_name]
                current["brand"] = printer["brand"]
                current["returnMonths"] = float(round_decimal(printer["return_months"]))
                current["machineCost"] = float(round_decimal(printer["machine_cost"]))
                current["workHoursPerDay"] = float(round_decimal(printer["work_hours_per_day"]))
                current["workingDaysPerMonth"] = float(round_decimal(printer["working_days_per_month"]))
                current["powerKw"] = float(round_decimal(printer["power_kw"]))
                current["usageLevel"] = printer["usage_level"]
                current["failureRate"] = float(round_decimal(printer["failure_rate"]))
                current["updatedAtUtc"] = SEED_TIMESTAMP

        for filament_name, filament in context.filaments.items():
            if filament_name not in filament_index:
                filament_index[filament_name] = {
                    "id": deterministic_uuid("filament", filament_name),
                    "name": filament_name,
                    "brand": filament["brand"],
                    "description": filament["description"],
                    "spoolWeightKg": float(round_decimal(filament["spool_weight_kg"])),
                    "spoolLengthMeters": float(round_decimal(filament["spool_length_meters"])),
                    "costBRL": float(round_decimal(filament["cost_brl"])),
                    "createdAtUtc": SEED_TIMESTAMP,
                    "updatedAtUtc": SEED_TIMESTAMP,
                }
            else:
                current = filament_index[filament_name]
                current["brand"] = filament["brand"]
                current["description"] = filament["description"]
                current["spoolWeightKg"] = float(round_decimal(filament["spool_weight_kg"]))
                current["spoolLengthMeters"] = float(round_decimal(filament["spool_length_meters"]))
                current["costBRL"] = float(round_decimal(filament["cost_brl"]))
                current["updatedAtUtc"] = SEED_TIMESTAMP

            cost_per_unit = round_decimal(filament["cost_brl"] / max(Decimal("1"), filament["spool_weight_kg"] * Decimal("1000")))
            if filament_name not in supply_index:
                supply_index[filament_name] = {
                    "id": deterministic_uuid("supply", filament_name),
                    "name": filament_name,
                    "unit": "g",
                    "costPerUnit": float(cost_per_unit),
                    "stockQuantity": float(STOCK_QUANTITY),
                    "minimumStock": float(MINIMUM_STOCK),
                    "notes": f"Filamento {filament['brand']} extraido da planilha",
                    "createdAtUtc": SEED_TIMESTAMP,
                    "updatedAtUtc": SEED_TIMESTAMP,
                }
            else:
                current = supply_index[filament_name]
                current["costPerUnit"] = float(cost_per_unit)
                current["notes"] = f"Filamento {filament['brand']} extraido da planilha"
                current["updatedAtUtc"] = SEED_TIMESTAMP

        for marketplace_name, marketplace in context.marketplaces.items():
            if marketplace_name not in marketplace_index:
                marketplace_index[marketplace_name] = {
                    "id": deterministic_uuid("marketplace", marketplace_name),
                    "name": marketplace_name,
                    "fixedFee": float(round_decimal(marketplace["fixed_fee"])),
                    "percentageFee": float(round_decimal(marketplace["percentage_fee"], "0.0001")),
                    "createdAtUtc": SEED_TIMESTAMP,
                    "updatedAtUtc": SEED_TIMESTAMP,
                }
            else:
                current = marketplace_index[marketplace_name]
                current["fixedFee"] = float(round_decimal(marketplace["fixed_fee"]))
                current["percentageFee"] = float(round_decimal(marketplace["percentage_fee"], "0.0001"))
                current["updatedAtUtc"] = SEED_TIMESTAMP

        for row in range(6, list_sheet.max_row + 1):
            product_name = list_sheet[f"B{row}"].value
            if not product_name or str(list_sheet[f"D{row}"].value or "").strip().lower() != "total":
                continue

            normalized_product_name = normalize_product_name(product_name)
            key = (context.category_name, normalized_product_name)
            position = product_index.get(key)
            if position is None:
                products.append(build_product_seed(context.category_name, normalized_product_name))
                position = len(products) - 1
                product_index[key] = position
                added_products += 1

            original_finishing = products[position].get("finishingPercentage")
            updated = recalculate_product(products[position], row, list_sheet, context)
            products[position] = updated
            updated_products += 1
            if original_finishing != updated["finishingPercentage"]:
                normalized_finishing += 1

    text = replace_json_constant(text, "CATEGORIES_JSON", list(category_index.values()))
    text = replace_json_constant(text, "PRINTERS_JSON", list(printer_index.values()))
    text = replace_json_constant(text, "FILAMENTS_JSON", list(filament_index.values()))
    text = replace_json_constant(text, "MARKETPLACES_JSON", list(marketplace_index.values()))
    text = replace_json_constant(text, "SUPPLIES_JSON", list(supply_index.values()))
    text = replace_json_constant(text, "PRODUCTS_JSON", products)
    SEED_FILE.write_text(text, encoding="utf-8")

    print(
        json.dumps(
            {
                "updatedProducts": updated_products,
                "addedProducts": added_products,
                "categories": len(category_index),
                "normalizedFinishingProducts": normalized_finishing,
                "printers": len(printer_index),
                "filaments": len(filament_index),
                "marketplaces": len(marketplace_index),
                "supplies": len(supply_index),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()